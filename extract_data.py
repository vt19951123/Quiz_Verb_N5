#!/usr/bin/env python3
"""
extract_data.py - Parse Bang_Chia_Dong_Tu_N5.xlsx and generate data.json
"""

import json
import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "Bang_Chia_Dong_Tu_N5.xlsx")
TXT_FILE = os.path.join(SCRIPT_DIR, "The_Ba_the_Ta.txt")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "data.json")

# Vietnamese translations for example sentences
EXAMPLE_TRANSLATIONS = {
    "会う": ["Gặp bạn bè.", "Không gặp được giám đốc."],
    "遊ぶ": ["Đã chơi ở công viên.", "Hôm nay không thể chơi được."],
    "洗う": ["Rửa tay.", "Đã không rửa bát."],
    "ある": ["Trên bàn có quyển sách.", "Đã không có tiền."],
    "歩く": ["Đã đi bộ đến nhà ga.", "Chân đau nên không thể đi bộ."],
    "言う": ["Nói \"Xin chào\".", "Anh ấy đã không nói gì."],
    "行く": ["Đi đến trường.", "Ngày mai không đi."],
    "歌う": ["Hát bài hát.", "Đã không hát karaoke."],
    "売る": ["Bán sách cũ.", "Không bán được."],
    "置く": ["Đặt sách lên bàn.", "Đã đặt chìa khóa ở đâu?"],
    "送る": ["Gửi thư.", "Đã không gửi email."],
    "押す": ["Nhấn nút.", "Đã nhấn nút khẩn cấp."],
    "思う": ["Tôi nghĩ là đúng.", "Đã không nghĩ đến điều đó."],
    "泳ぐ": ["Bơi ở biển.", "Không thể bơi được."],
    "終わる": ["Công việc kết thúc.", "Vẫn chưa kết thúc."],
    "買う": ["Mua sách.", "Đã không mua quần áo."],
    "返す": ["Trả lại sách.", "Chưa trả lại tiền."],
    "帰る": ["Về nhà.", "Hôm qua đã về muộn."],
    "かかる": ["Mất 1 tiếng.", "Đã không mất thời gian."],
    "書く": ["Viết thư.", "Đã không viết nhật ký."],
    "貸す": ["Cho mượn sách.", "Đã không cho mượn tiền."],
    "被る": ["Đội mũ.", "Hôm qua đã không đội mũ."],
    "借りる": ["Mượn sách.", "Không thể mượn được."],
    "聞く": ["Nghe nhạc.", "Đã hỏi thầy giáo."],
    "切る": ["Cắt giấy.", "Đã cắt tóc."],
    "消す": ["Tắt đèn.", "Đã không tắt TV."],
    "答える": ["Trả lời câu hỏi.", "Đã không trả lời được."],
    "困る": ["Gặp khó khăn.", "Đã gặp khó khăn vì không có tiền."],
    "咲く": ["Hoa nở.", "Hoa chưa nở."],
    "差す": ["Che ô.", "Trời đẹp nên không cần che ô."],
    "死ぬ": ["Cá đã chết.", "Không chết được."],
    "閉まる": ["Cửa đóng.", "Cửa không đóng."],
    "閉める": ["Đóng cửa.", "Đã không đóng cửa sổ."],
    "知る": ["Biết câu trả lời.", "Không biết tên anh ấy."],
    "吸う": ["Hút thuốc.", "Đã bỏ hút thuốc."],
    "住む": ["Sống ở Tokyo.", "Không sống ở đây."],
    "座る": ["Ngồi xuống ghế.", "Không có chỗ ngồi."],
    "出す": ["Lấy ví ra.", "Đã nộp bài tập."],
    "立つ": ["Đứng dậy.", "Đã đứng 2 tiếng."],
    "頼む": ["Nhờ bạn giúp.", "Đã đặt bia."],
    "食べる": ["Ăn cơm.", "Sáng nay đã không ăn gì."],
    "違う": ["Câu trả lời sai rồi.", "Đây không phải cái đó."],
    "使う": ["Dùng máy tính.", "Đã không dùng điện thoại."],
    "疲れる": ["Mệt rồi.", "Hôm qua rất mệt."],
    "着く": ["Đến nhà ga.", "Vẫn chưa đến."],
    "作る": ["Làm cơm.", "Đã không làm bánh."],
    "つける": ["Bật đèn.", "Đã không bật máy lạnh."],
    "勤める": ["Làm việc ở ngân hàng.", "Trước đây làm ở bệnh viện."],
    "出る": ["Ra khỏi nhà.", "Đã ra ngoài lúc 7 giờ."],
    "飛ぶ": ["Máy bay bay.", "Không bay được."],
    "止まる": ["Xe dừng lại.", "Đồng hồ đã dừng."],
    "撮る": ["Chụp ảnh.", "Đã không quay video."],
    "取る": ["Lấy muối.", "Đã lấy kỳ nghỉ."],
    "鳴く": ["Chó sủa.", "Mèo không kêu."],
    "なくす": ["Làm mất ví.", "Đã làm mất chìa khóa."],
    "並ぶ": ["Xếp hàng.", "Đã xếp hàng 1 tiếng."],
    "並べる": ["Xếp sách lên kệ.", "Đã bày đồ ăn lên bàn."],
    "習う": ["Học tiếng Nhật.", "Đã học piano."],
    "なる": ["Trở thành bác sĩ.", "Trời đã tối."],
    "脱ぐ": ["Cởi giày.", "Đã không cởi áo khoác."],
    "登る": ["Leo núi.", "Đã không leo được."],
    "飲む": ["Uống nước.", "Đã không uống thuốc."],
    "乗る": ["Lên xe buýt.", "Đã không lên tàu."],
    "入る": ["Vào phòng.", "Không được vào."],
    "履く": ["Mang giày.", "Đã mang giày mới."],
    "始まる": ["Lớp học bắt đầu.", "Phim chưa bắt đầu."],
    "走る": ["Chạy ở công viên.", "Đã không chạy được."],
    "働く": ["Làm việc ở công ty.", "Hôm qua đã không làm việc."],
    "話す": ["Nói chuyện với bạn.", "Đã không nói với ai."],
    "貼る": ["Dán tem.", "Đã không dán ảnh."],
    "晴れる": ["Trời nắng.", "Ngày mai có lẽ sẽ nắng."],
    "引く": ["Kéo cửa.", "Bị cảm."],
    "弾く": ["Chơi piano.", "Không thể chơi guitar."],
    "吹く": ["Gió thổi.", "Đã thổi nến."],
    "降る": ["Trời mưa.", "Tuyết không rơi."],
    "曲がる": ["Rẽ phải.", "Rẽ trái ở ngã tư."],
    "待つ": ["Đợi bạn.", "Đã đợi 30 phút."],
    "磨く": ["Đánh răng.", "Đã không đánh răng."],
    "見る": ["Xem TV.", "Đã không xem phim."],
    "見せる": ["Cho xem ảnh.", "Đã không cho xem."],
    "持つ": ["Cầm túi.", "Không có ô."],
    "休む": ["Nghỉ ngơi.", "Hôm qua đã nghỉ học."],
    "やる": ["Làm bài tập.", "Đã không làm gì cả."],
    "呼ぶ": ["Gọi taxi.", "Đã gọi tên bạn."],
    "読む": ["Đọc sách.", "Đã không đọc báo."],
    "分かる": ["Hiểu rồi.", "Không hiểu tiếng Nhật."],
    "忘れる": ["Quên mang ô.", "Đã quên mật khẩu."],
    "渡る": ["Sang đường.", "Đã qua cầu."],
    "渡す": ["Đưa tài liệu.", "Đã đưa quà."],
    "いる": ["Có mèo ở đây.", "Không có ai."],
    "起きる": ["Thức dậy lúc 7 giờ.", "Sáng nay đã không dậy sớm."],
    "教える": ["Dạy tiếng Nhật.", "Đã không dạy toán."],
    "覚える": ["Nhớ từ vựng.", "Không nhớ được."],
    "降りる": ["Xuống xe buýt.", "Đã xuống sai trạm."],
    "開ける": ["Mở cửa.", "Đã không mở cửa sổ."],
    "あげる": ["Cho quà.", "Đã tặng hoa."],
    "浴びる": ["Tắm vòi sen.", "Sáng nay đã không tắm."],
    "来る": ["Bạn đến.", "Hôm qua đã không đến."],
    "する": ["Làm bài tập.", "Đã không làm gì."],
    "散歩する": ["Đi dạo ở công viên.", "Hôm qua đã không đi dạo."],
    "勉強する": ["Học tiếng Nhật.", "Đã không học."],
    "結婚する": ["Kết hôn năm ngoái.", "Chưa kết hôn."],
    "買い物する": ["Mua sắm ở siêu thị.", "Đã không mua sắm được."],
    "掛かる": ["Mất 10 phút.", "Đã không mất nhiều thời gian."],
    "頑張る": ["Cố gắng học.", "Hãy cố gắng lên."],
    "触る": ["Chạm vào tranh.", "Xin đừng chạm."],
    "直す": ["Sửa xe đạp.", "Đã không sửa được."],
    "泣く": ["Em bé khóc.", "Đã không khóc."],
    "払う": ["Trả tiền.", "Đã không trả tiền."],
    "貰う": ["Nhận đồng hồ.", "Đã được nhận quà."],
    "上げる": ["Giơ tay lên.", "Đã cho bạn quà."],
    "集める": ["Thu thập tiền.", "Sưu tầm tem."],
    "あり得る": ["Điều đó có thể xảy ra.", "Điều đó không thể xảy ra."],
    "居る": ["Có con chó.", "Không có ai ở nhà."],
    "入れる": ["Cho sách vào.", "Đã cho đường vào cà phê."],
    "変える": ["Thay đổi lịch trình.", "Đã không thay đổi kế hoạch."],
    "掛ける": ["Gọi điện thoại.", "Đã treo áo lên."],
    "消える": ["Đèn tắt.", "Chữ đã biến mất."],
    "着る": ["Mặc áo khoác.", "Hôm qua đã không mặc áo ấm."],
    "調べる": ["Tra cứu từ vựng.", "Đã tra trên internet."],
    "付ける": ["Bật đèn.", "Đã không bật máy lạnh."],
    "出来る": ["Có thể nấu ăn.", "Cơm đã nấu xong."],
    "出掛ける": ["Đi mua sắm.", "Hôm qua đã đi ra ngoài."],
    "止める": ["Dừng xe.", "Đã dừng máy."],
    "寝る": ["Ngủ lúc 11 giờ.", "Hôm qua đã ngủ muộn."],
    "始める": ["Bắt đầu cuộc họp.", "Đã bắt đầu học."],
    "見える": ["Nhìn thấy biển.", "Không nhìn thấy gì."],
    "辞める": ["Nghỉ việc ở công ty.", "Đã bỏ hút thuốc."],
}


def parse_ba_ta_file():
    """Parse The_Ba_the_Ta.txt and return dict keyed by kanji"""
    ba_ta_data = {}
    with open(TXT_FILE, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    # Skip header line
    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue
        parts = line.split(',')
        if len(parts) < 5:
            continue
        # Column 0: "会う - あう - (gặp)"
        verb_info = parts[0].strip()
        verb_parts = verb_info.split(' - ')
        if len(verb_parts) < 2:
            continue
        kanji = verb_parts[0].strip()
        
        ba_form = parts[1].strip()
        ba_example = parts[2].strip()
        ta_form = parts[3].strip()
        ta_example = parts[4].strip() if len(parts) > 4 else ""
        
        ba_ta_data[kanji] = {
            "ba_form": ba_form,
            "ba_example": ba_example,
            "ta_form": ta_form,
            "ta_example": ta_example
        }
    return ba_ta_data


def parse_verb_cell(cell_value):
    """Parse the first column cell: 'Kanji\\nReading\\n(meaning)'"""
    if not cell_value:
        return None, None, None
    lines = cell_value.strip().split('\n')
    kanji = lines[0].strip() if len(lines) > 0 else ""
    reading = lines[1].strip() if len(lines) > 1 else kanji
    meaning = ""
    if len(lines) > 2:
        m = lines[2].strip()
        if m.startswith('(') and m.endswith(')'):
            meaning = m[1:-1]
        else:
            meaning = m
    return kanji, reading, meaning


def parse_examples(jp_cell, reading_cell):
    """Parse example sentences from cells with ・ prefix and \\n separator"""
    examples = []
    if not jp_cell:
        return examples
    jp_lines = [l.strip().lstrip('・').strip() for l in jp_cell.strip().split('\n') if l.strip()]
    rd_lines = []
    if reading_cell:
        rd_lines = [l.strip().lstrip('・').strip() for l in reading_cell.strip().split('\n') if l.strip()]
    for i, jp in enumerate(jp_lines):
        rd = rd_lines[i] if i < len(rd_lines) else ""
        examples.append({"japanese": jp, "reading": rd})
    return examples


def extract_verbs():
    """Extract all verb data from the Excel file"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    verbs = []
    verb_id = 0
    row = 3  # Data starts at row 3

    while row <= ws.max_row:
        cell_val = ws.cell(row=row, column=1).value
        if cell_val is None:
            row += 1
            continue

        kanji, reading, meaning = parse_verb_cell(cell_val)
        if not kanji:
            row += 1
            continue

        verb_id += 1
        group = ws.cell(row=row, column=2).value

        # Parse 6 rows of conjugations
        conjugations = {"present": {}, "past": {}}
        potential = {"present": {}, "past": {}}

        tense_map = {"hiện tại": "present", "quá khứ": "past"}
        form_map = {"Khẳng định": "affirmative", "phủ định": "negative", "nghi vấn": "question"}

        for i in range(6):
            r = row + i
            if r > ws.max_row:
                break
            tense_raw = ws.cell(row=r, column=3).value
            form_raw = ws.cell(row=r, column=4).value

            # Determine current tense
            if tense_raw and tense_raw in tense_map:
                current_tense = tense_map[tense_raw]
            
            if not form_raw or form_raw not in form_map:
                continue

            form_key = form_map[form_raw]
            polite_kn = ws.cell(row=r, column=5).value or "-"
            plain_kn = ws.cell(row=r, column=6).value or "-"
            polite_k = ws.cell(row=r, column=7).value or "-"
            plain_k = ws.cell(row=r, column=8).value or "-"

            conjugations[current_tense][form_key] = {
                "polite": polite_kn,
                "plain": plain_kn
            }
            potential[current_tense][form_key] = {
                "polite": polite_k,
                "plain": plain_k
            }

        # Parse examples
        examples = parse_examples(
            ws.cell(row=row, column=9).value,
            ws.cell(row=row, column=10).value
        )

        # Add Vietnamese translations
        vn_translations = EXAMPLE_TRANSLATIONS.get(kanji, [])
        for i, ex in enumerate(examples):
            ex["vietnamese"] = vn_translations[i] if i < len(vn_translations) else ""

        verb = {
            "id": verb_id,
            "kanji": kanji,
            "reading": reading,
            "meaning": meaning,
            "group": group,
            "conjugations": conjugations,
            "potential": potential,
            "examples": examples
        }
        verbs.append(verb)
        row += 6  # Jump to next verb

    # Merge ba/ta data
    ba_ta_data = parse_ba_ta_file()
    merged = 0
    for verb in verbs:
        bt = ba_ta_data.get(verb["kanji"])
        if bt:
            verb["ba_form"] = bt["ba_form"]
            verb["ba_example"] = bt["ba_example"]
            verb["ta_form"] = bt["ta_form"]
            verb["ta_example"] = bt["ta_example"]
            merged += 1
        else:
            verb["ba_form"] = ""
            verb["ba_example"] = ""
            verb["ta_form"] = ""
            verb["ta_example"] = ""
    print(f"  -> Merged ba/ta data for {merged}/{len(verbs)} verbs")

    return verbs


def main():
    print("=== Extract Dong Tu N5 ===")
    verbs = extract_verbs()
    print(f"  -> Extracted {len(verbs)} verbs")

    # Validate
    for v in verbs:
        if not v["meaning"]:
            print(f"  [WARN] Missing meaning: {v['kanji']}")
        if not v["examples"]:
            print(f"  [WARN] No examples: {v['kanji']}")
        vn = EXAMPLE_TRANSLATIONS.get(v["kanji"])
        if not vn:
            print(f"  [WARN] No VN translation: {v['kanji']}")

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(verbs, f, ensure_ascii=False, indent=2)
    print(f"  -> Saved to {OUTPUT_FILE}")

    # Print sample
    if verbs:
        print(f"\n  Sample: {verbs[0]['kanji']} ({verbs[0]['reading']}) = {verbs[0]['meaning']}")
        print(f"  Group: {verbs[0]['group']}")
        print(f"  Examples: {len(verbs[0]['examples'])}")


if __name__ == "__main__":
    main()
